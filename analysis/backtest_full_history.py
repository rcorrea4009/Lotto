#!/usr/bin/env python3
"""
Empirical test: can ANY model (frequency-weighted, recency-weighted, or a
trained logistic-regression scorer) beat chance at predicting NZ Lotto main
numbers, using the FULL official history since draw #1 (1987)?

Everything here is walk-forward / no-lookahead: a prediction for draw i is
built ONLY from draws strictly before i. This is not a demonstration of a
predictive system -- it is the opposite: a rigorous attempt to find an edge,
reported honestly whichever way it comes out.

Run: python3 backtest_full_history.py
"""
import json
import statistics
from collections import Counter
from pathlib import Path

import numpy as np
import openpyxl
from scipy import stats as sstats
from sklearn.linear_model import LogisticRegression

UPLOAD_XLSX = "/root/.claude/uploads/9eb1be5a-7fde-56e1-943d-b276dc61466c/7eb16d4e-2cd0eeb28fa711f196a8b6742211afef.xlsx"
POOL = list(range(1, 41))
ALPHA = 1.0
RECENCY_WEIGHT = 0.35


def load_full_history():
    wb = openpyxl.load_workbook(UPLOAD_XLSX, data_only=True)
    ws = wb["Lotto Powerball"]
    draws = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 9)]
        date, draw_no, n1, n2, n3, n4, n5, n6 = vals
        if draw_no is None:
            continue
        nums = sorted(int(x) for x in [n1, n2, n3, n4, n5, n6])
        draws.append({"draw": int(draw_no), "date": str(date), "numbers": nums})
    draws.sort(key=lambda d: d["draw"])  # oldest first, chronological
    return draws


def chi_square_test(draws):
    counts = Counter()
    for d in draws:
        for n in d["numbers"]:
            counts[n] += 1
    total_balls = len(draws) * 6
    expected = total_balls / 40
    observed = [counts[n] for n in POOL]
    chi2, p = sstats.chisquare(observed, f_exp=[expected] * 40)
    deviations = sorted(
        ((n, counts[n], counts[n] - expected) for n in POOL), key=lambda x: -abs(x[2])
    )
    return {
        "n_draws": len(draws),
        "total_balls_drawn": total_balls,
        "expected_count_per_number": round(expected, 1),
        "chi2_statistic": round(chi2, 2),
        "degrees_of_freedom": 39,
        "p_value": float(p),
        "verdict": (
            "NOT statistically significant (p > 0.05) -- no evidence any number "
            "deviates from a fair 1/40 chance across the full history"
            if p > 0.05 else "statistically significant deviation -- see note"
        ),
        "top_deviations": [
            {"number": n, "observed": c, "expected": round(expected, 1), "diff": round(diff, 1)}
            for n, c, diff in deviations[:5]
        ],
    }


class IncrementalFeatures:
    """Precomputes, in one O(N) pass, everything needed to score all 40
    numbers as of just-before any draw index i -- so the backtest loop can
    look features up in O(1) instead of rescanning history each step."""

    def __init__(self, draws):
        n = len(draws)
        self.n = n
        # prefix[num][i] = count of `num` in draws[0..i-1]
        self.prefix = {num: np.zeros(n + 1, dtype=np.int32) for num in POOL}
        # gap_at[num][i] = draws since `num` last appeared, as of just before draw i
        self.gap_at = {num: np.zeros(n, dtype=np.int32) for num in POOL}
        last_seen = {num: -1 for num in POOL}
        for i, d in enumerate(draws):
            for num in POOL:
                self.prefix[num][i + 1] = self.prefix[num][i]
                self.gap_at[num][i] = i if last_seen[num] == -1 else i - 1 - last_seen[num]
            for num in d["numbers"]:
                self.prefix[num][i + 1] += 1
                last_seen[num] = i

    def freq_full(self, num, i):
        return int(self.prefix[num][i])

    def freq_window(self, num, i, w):
        lo = max(0, i - w)
        return int(self.prefix[num][i] - self.prefix[num][lo])

    def gap(self, num, i):
        return int(self.gap_at[num][i])

    def blended_probs(self, i, window):
        n_full = i
        n_win = min(i, window) if window else i
        k = 40
        probs = {}
        for num in POOL:
            long_rate = (self.freq_full(num, i) + ALPHA) / (n_full * 6 + ALPHA * k) if n_full else 1 / k
            recent_rate = (self.freq_window(num, i, window) + ALPHA) / (n_win * 6 + ALPHA * k) if n_win else 1 / k
            probs[num] = (1 - RECENCY_WEIGHT) * long_rate + RECENCY_WEIGHT * recent_rate
        total = sum(probs.values())
        return {num: p / total for num, p in probs.items()}

    def feature_row(self, num, i):
        n = max(i, 1)
        return [
            self.freq_full(num, i) / n,
            self.freq_window(num, i, 10) / 10.0,
            self.freq_window(num, i, 30) / 30.0,
            min(self.gap(num, i), 60) / 60.0,
        ]


def top6(probs):
    return sorted(sorted(probs, key=lambda n: -probs[n])[:6])


def weighted_sample_6(probs, rng):
    pool = list(probs.keys())
    weights = [probs[n] for n in pool]
    picks = []
    pool, weights = pool[:], weights[:]
    for _ in range(6):
        total = sum(weights)
        r = rng.uniform(0, total)
        upto = 0
        for idx, w in enumerate(weights):
            upto += w
            if upto >= r:
                picks.append(pool.pop(idx))
                weights.pop(idx)
                break
    return picks


def backtest(draws, feats, warmup=200, retrain_every=200):
    rng = np.random.default_rng(42)
    results = {
        "top6_freq_full_history": [], "top6_freq_recent50": [], "top6_freq_recent20": [],
        "weighted_sample_freq": [], "random_uniform": [], "top6_logreg": [],
    }
    n = len(draws)

    logreg_X, logreg_y = [], []
    clf = None

    for i in range(warmup, n):
        actual = set(draws[i]["numbers"])

        p_full = feats.blended_probs(i, window=i)
        results["top6_freq_full_history"].append(len(set(top6(p_full)) & actual))

        p_r50 = feats.blended_probs(i, window=50)
        results["top6_freq_recent50"].append(len(set(top6(p_r50)) & actual))

        p_r20 = feats.blended_probs(i, window=20)
        results["top6_freq_recent20"].append(len(set(top6(p_r20)) & actual))

        ws = weighted_sample_6(p_full, rng)
        results["weighted_sample_freq"].append(len(set(ws) & actual))

        rnd = rng.choice(POOL, size=6, replace=False)
        results["random_uniform"].append(len(set(rnd.tolist()) & actual))

        # logistic regression: predict with the model trained on everything
        # strictly before this point, then retrain periodically
        if clf is not None:
            X_now = np.array([feats.feature_row(num, i) for num in POOL])
            proba = clf.predict_proba(X_now)[:, 1]
            picks = sorted(sorted(range(40), key=lambda idx: -proba[idx])[:6])
            picks = [POOL[idx] for idx in picks]
            results["top6_logreg"].append(len(set(picks) & actual))

        for num in POOL:
            logreg_X.append(feats.feature_row(num, i))
            logreg_y.append(1 if num in actual else 0)

        if (i - warmup) % retrain_every == 0 and len(logreg_X) > 400:
            clf = LogisticRegression(max_iter=300, class_weight="balanced")
            clf.fit(np.array(logreg_X), np.array(logreg_y))

    summary = {}
    for key, vals in results.items():
        if not vals:
            continue
        mean = statistics.mean(vals)
        se = statistics.pstdev(vals) / (len(vals) ** 0.5) if len(vals) > 1 else 0
        summary[key] = {
            "n_predictions": len(vals),
            "avg_matches_of_6": round(mean, 4),
            "std_error": round(se, 4),
            "95pct_ci": [round(mean - 1.96 * se, 4), round(mean + 1.96 * se, 4)],
        }
    summary["theoretical_expected_if_pure_chance"] = round(6 * 6 / 40, 4)
    return summary


def main():
    print("Loading full official history from the uploaded workbook...")
    draws = load_full_history()
    print(f"Loaded {len(draws)} draws: #{draws[0]['draw']} ({draws[0]['date']}) "
          f"to #{draws[-1]['draw']} ({draws[-1]['date']})")
    print()

    print("=== TEST 1: Chi-square uniformity test, full history ===")
    chi = chi_square_test(draws)
    print(json.dumps(chi, indent=2))
    print()

    print("=== TEST 2: Walk-forward backtest (strict no-lookahead) ===")
    print(f"Predicting draws #{draws[200]['draw']} through #{draws[-1]['draw']} "
          f"({len(draws) - 200} predictions per model), each built only from "
          "draws strictly before it.")
    feats = IncrementalFeatures(draws)
    bt = backtest(draws, feats, warmup=200, retrain_every=200)
    print(json.dumps(bt, indent=2))

    out = {"chi_square_test": chi, "backtest": bt, "n_draws_used": len(draws)}
    Path("analysis/backtest_report.json").write_text(json.dumps(out, indent=2))
    print("\nWrote analysis/backtest_report.json")


if __name__ == "__main__":
    main()
